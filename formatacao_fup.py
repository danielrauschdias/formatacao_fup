import pandas as pd;
import numpy as np;
from pathlib import Path;
from datetime import datetime;
from datetime import date;
import locale;
import os;
import shutil
from openpyxl import load_workbook;
from openpyxl.worksheet.datavalidation import DataValidation;
from openpyxl.worksheet.table import Table, TableStyleInfo;
from openpyxl.styles import PatternFill;
from openpyxl.styles import Color;
from openpyxl.styles import Font;
from openpyxl.styles import Alignment;
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder;
from openpyxl.utils import get_column_letter;
from openpyxl.comments import Comment;
import tkinter as tk;
from tkinter import filedialog;
from tkinter import ttk;
from tkinter import messagebox;
from PIL import ImageTk, Image;
import warnings;

locale.setlocale(locale.LC_TIME, "pt_BR.UTF-8")

#Cria o backup da pasta de liberação antes de formatar a planilha.
def criarBackupPastaLiberacao():
    caminho_Pasta_Liberacao = ent_caminho_pasta_liberacao.get()
    caminho_Local_Pasta_Liberacao = Path(caminho_Pasta_Liberacao).parent
    (caminho_Local_Pasta_Liberacao/"Backup").mkdir(parents=True, exist_ok=True)
    shutil.copy(caminho_Pasta_Liberacao, caminho_Local_Pasta_Liberacao/"Backup"/f"{Path(caminho_Pasta_Liberacao).name}")
    

def formatar():
    dataHoje = date.today().strftime("%d-%m-%Y");

    listaTopVazios = []
    listaTopExist = []
    listaTopicos = []

    #Verifica quais tópicos existem, se existir, adiciona o DataFrame à lista listaTopicos, o valor "True" à lista listaTopExist e "False" à lista listaTopVazios, caso contrário, adiciona uma string vazia à lista listaTopicos, "False" à lista listaTopExist e "True" à lista listaTopVazios.
    caminho_topicos = ent_caminho_topicos.get()
    if caminho_topicos == "":
        messagebox.showerror("Caminho inválido", "Favor informar o caminho das planilhas dos tópicos.");
        return
    for i in range(1, 3):
        tabelaTopCaminho = caminho_topicos + f"\{dataHoje} T{i}.xlsx"
        if os.path.isfile(tabelaTopCaminho):
            listaTopicos.append(pd.read_excel(tabelaTopCaminho))
            listaTopExist.append(True);
            listaTopVazios.append(False);
        else:
            listaTopicos.append("");
            listaTopExist.append(False);
            listaTopVazios.append(True);
    
    #Verifica se pelo menos um dos tópicos existe, se não existir, exibe uma mensagem de erro e retorna para a tela inicial.
    if not listaTopExist[0] and not listaTopExist[1]:
        messagebox.showerror("Tópicos não encontrados", 
                             "Não foi encontrada a pasta de trabalho de nenhum dos dois tópicos no caminho " +  caminho_topicos +
                             " favor colocar a de pelo menos um deles na pasta e tentar novamente.")
        return

    #Verifica se foi digitado algum valor incorreto nos campos de dias de análise, se sim, exibe uma mensagem de erro e retorna para a tela inicial.
    try:
        diasTransf = int(ent_entregaInbound.get());
        diasCompra = int(ent_ordemPlanejada.get());
        diasCompraCDB = int(ent_ordemPlanejadaCDB.get());
    except ValueError:
        messagebox.showerror("Valor inválido", 
                                "Foi informado um valor inválido para os dias de análise, são permitidos apenas valores numéricos inteiros, favor informar novamente.");
        return

    #Se os valores de dias de análise forem negativos, exibe uma mensagem de erro e retorna para a tela inicial.
    if (diasTransf < 0 or diasCompra < 0):
        messagebox.showerror("Valor negativo", "Foi informado um valor negativo nos dias de análise, favor informar somente valores positivos.");
        return
    
    #Se foi digitado um número no lugar de texto para a unidade de negócio, exibe uma mensagem de erro e retorna para a tela inicial.
    undNegoc = combo_undNegoc.get();
    if (undNegoc.isdigit()):
        messagebox.showerror("Valor inválido", "Foi informado um dígito para o nome da unidade de négócio, favor informar apenas palavras.");
        return

    #Se a unidade de negócio for Helicopteros e o valor de dias de análise de compra CDB for negativo, exibe uma mensagem de erro e retorna para a tela inicial.
    if (undNegoc == "Helicopteros" and diasCompraCDB < 0):
        messagebox.showerror("Valor inválido", "Foi informado um valor negativo para os dias de análise de compra CDB, favor informar somente valores positivos.");
        return

    wbLiberacaoCaminho = ent_caminho_pasta_liberacao.get();
    caminho_Itens_Restricao = ent_caminho_pasta_restricao.get()
    #Abre o arquivo de filtros, apaga os valores anteriores e insere os valores que foram utilizados na última formatação.
    modoAbertura = "w"
    espaco = " \n"
    arqFiltros = open(caminho_Filtros, modoAbertura)
    arqFiltros.writelines([caminho_topicos + espaco, wbLiberacaoCaminho + espaco, caminho_Itens_Restricao + espaco, undNegoc + espaco, str(diasTransf) + espaco, str(diasCompra) + espaco, str(diasCompraCDB) + espaco])
    arqFiltros.close();

    #Define as datas de comparação para os dias de análise de transferência, compra e compra CDB.
    dataComparacaoTransf = datetime.now()+pd.Timedelta(days=diasTransf);
    dataComparacaoCompra = datetime.now()+pd.Timedelta(days=diasCompra);
    dataComparacaoCompraCDB = datetime.now()+pd.Timedelta(days=diasCompraCDB);

    dataHoje = date.today().strftime("%d-%m");

    wsTopico1Nome = dataHoje + " T1"
    wsTopico2Nome = dataHoje + " T2"

    wbLiberacao = load_workbook(wbLiberacaoCaminho);

    #Percorre todas as planilhas da pasta de trabalho da liberação para verificar se existe a planilha de relatório ou se já existe a planilha de liberação do dia, em ambos os casos, exibe uma mensagem de erro e retorna para a tela inicial.
    for ws_name in wbLiberacao.sheetnames:
        if ws_name == "Relatório":
            wbLiberacao.close()
            messagebox.showerror("Planilha de Relatório", "Favor apagar a planilha de relatório antes de formatar a planilha de liberação.")
            return
        elif ws_name == dataHoje[0:4]:
            wbLiberacao.close()
            messagebox.showerror("Planilha de liberação do dia já existente", "A planilha de liberação do dia já existe, favor apagar antes de tentar formatar novamente.")
            return

    numLinhasT1 = 0;
    numLinhasT2 = 0;

    #Função para filtrar o tópico 1
    def filtraTopico1(wbLiberacaoCaminho, wsTopico1Nome, listaTopicos):
        nomesColunas = [];

        for coluna in listaTopicos[0].columns.array:
            nomesColunas.append(coluna);
        
        listaTopicos[0] = listaTopicos[0].drop_duplicates();

        arrayTop1Extrac = np.array(listaTopicos[0]);
       
        colUndNegoc = 1;
        
        arrayFiltroUndNegoc = np.array([registro for registro in arrayTop1Extrac if registro[colUndNegoc] == undNegoc]);

        topVazio = False;

        if arrayFiltroUndNegoc.size == 0:

            topVazio = True;

            tabelaTop1Filtrada = pd.DataFrame(index=range(1), columns=nomesColunas);

            tabelaTop1Filtrada["PN&PREFIXO&QTY"] = "";

            tabelaTop1Filtrada["Problema Raiz"] = "";

            tabelaTop1Filtrada["Ação"] = "";

            tabelaTop1Filtrada["OBS"] = "";

            tabelaTop1Filtrada["PN s/C"] = "";

            tabelaTop1Filtrada["Condição"] = "";
            
            tabelaTop1Filtrada["Desc."] = "";
        
        else:
            tabelaTop1Filtrada = pd.DataFrame(arrayFiltroUndNegoc, columns=nomesColunas);

            tabelaTop1Filtrada["PN&PREFIXO&QTY"] = tabelaTop1Filtrada["PN"] + tabelaTop1Filtrada["PREFIXO"].apply(str) + tabelaTop1Filtrada["QTDE. PENDENTE"].apply(int).apply(str);

            tabelaTop1Filtrada["Problema Raiz"] = "";

            tabelaTop1Filtrada["Ação"] = "";

            tabelaTop1Filtrada["OBS"] = "";

            tabelaTop1Filtrada["PN s/C"] = tabelaTop1Filtrada["PN"].str[:-2];

            tabelaTop1Filtrada["Condição"] = tabelaTop1Filtrada["PN"].str[-2:];
            
            tabelaTop1Filtrada["Desc."] = tabelaTop1Filtrada["DESCRIÇÃO"];
            
        with pd.ExcelWriter(wbLiberacaoCaminho, mode="a", if_sheet_exists="overlay") as writer:
            tabelaTop1Filtrada.to_excel(excel_writer=writer, sheet_name=wsTopico1Nome, index=False);

        return topVazio;

    #Função para filtrar o tópico 2
    def filtraTopico2(wbLiberacaoCaminho, wsTopico2Nome, listaTopicos):
        nomesColunas = [];

        for coluna in listaTopicos[1].columns.array:
            nomesColunas.append(coluna);
        
        listaTopicos[1].drop_duplicates(inplace=True);
            
        arrayTop2Extrac = np.array(listaTopicos[1]);

        colDtOrdemSug = 15;

        colTipoOrdem = 5;

        colUndNegoc = 3;

        colOrgOrigem = 6;

        colPrefixo = 12;

        tabelaFiltroUndNegoc = np.array([registro for registro in arrayTop2Extrac if registro[colUndNegoc] == undNegoc]);

        if undNegoc == "UOH":

            tabelaFiltroDataOrdemSugerida = np.array([registro for registro in tabelaFiltroUndNegoc if registro[colDtOrdemSug] <= dataComparacaoTransf and registro[colTipoOrdem] == "Entrega inbound planejada" or registro[colDtOrdemSug] <= dataComparacaoCompra and registro[colTipoOrdem] == "Ordem planejada" or registro[colOrgOrigem]=="CDB" and registro[colDtOrdemSug] <= dataComparacaoCompraCDB]);
        else:
            tabelaFiltroDataOrdemSugerida = np.array([registro for registro in tabelaFiltroUndNegoc if registro[colDtOrdemSug] <= dataComparacaoTransf and registro[colTipoOrdem] == "Entrega inbound planejada" or registro[colDtOrdemSug] <= dataComparacaoCompra and registro[colTipoOrdem] == "Ordem planejada"]);

        tabelaFiltroCD = np.array([registro for registro in tabelaFiltroDataOrdemSugerida if (registro[colOrgOrigem][:-1] != "CD" and registro[colTipoOrdem] == "Entrega inbound planejada") or registro[colTipoOrdem] == "Ordem planejada"]);

        tabelaFiltroPrefixo = np.array([registro for registro in tabelaFiltroCD if registro[colPrefixo] != "CONSUMO" and registro[colPrefixo] != "UNIFORME"]);

        for registro in tabelaFiltroPrefixo:
            if (type(registro[12]) != str):
                registro[12] = "";

        topVazio = False;

        if tabelaFiltroPrefixo.size == 0:
            topVazio = True

            tabelaTop2Filtrada = pd.DataFrame(index=range(1), columns=nomesColunas)

            tabelaTop2Filtrada["PN&PREFIXO&QTY"] = "";

            tabelaTop2Filtrada["Problema Raiz"] = "";

            tabelaTop2Filtrada["Ação"] = "";

            tabelaTop2Filtrada["OBS"] = "";

            tabelaTop2Filtrada["PN s/C"] = "";

            tabelaTop2Filtrada["Condição"] = "";
            
            tabelaTop2Filtrada["Desc."] = "";

        else:
            tabelaTop2Filtrada = pd.DataFrame(tabelaFiltroPrefixo, columns=nomesColunas);

            tabelaTop2Filtrada["PN&PREFIXO&QTY&TIPODEORDEM"] = tabelaTop2Filtrada["PN"] + tabelaTop2Filtrada["PREFIXO"].apply(str) + tabelaTop2Filtrada["QTDE"].apply(int).apply(str) + tabelaTop2Filtrada["TIPO ORDEM"];

            tabelaTop2Filtrada["Problema Raiz"] = "";

            tabelaTop2Filtrada["Ação"] = "";

            tabelaTop2Filtrada["OBS"] = "";

            tabelaTop2Filtrada["PN s/C"] = tabelaTop2Filtrada["PN"].str[:-2];

            tabelaTop2Filtrada["Condição"] = tabelaTop2Filtrada["PN"].str[-2:];
                
            tabelaTop2Filtrada["Desc."] = tabelaTop2Filtrada["DESCRIÇÃO"];

            tabelaTop2Filtrada = tabelaTop2Filtrada.sort_values("TIPO ORDEM");
        
        dictDocumentoCont = tabelaTop2Filtrada["NUM. DOCUMENTO"].value_counts().to_dict();

        dictDocumentoContFiltrado = {};

        dictDocumentoPrefixoVarios = {};

        dictDocumentoOsVarias = {};

        for documento in dictDocumentoCont:
            if dictDocumentoCont[documento] > 1:
                dictDocumentoContFiltrado.update({documento: dictDocumentoCont[documento]});
        
        for documento in dictDocumentoContFiltrado:
            listaPrefixos = [];
            listaOs = [];
            for linha in range(0, len(tabelaTop2Filtrada["NUM. DOCUMENTO"])):
                if tabelaTop2Filtrada.loc[linha, "NUM. DOCUMENTO"] == documento:
                    listaPrefixos.append(tabelaTop2Filtrada.loc[linha, "PREFIXO"]);
                    listaOs.append(tabelaTop2Filtrada.loc[linha, "OS"]);
            prefixoComp = listaPrefixos[0];
            for prefixo in listaPrefixos:
                if prefixo != prefixoComp:
                    dictDocumentoPrefixoVarios.update({documento: True});
                    break;
                else:
                    dictDocumentoPrefixoVarios.update({documento: False});
            osComp = listaOs[0];
            for os in listaOs:
                if os != osComp:
                    dictDocumentoOsVarias.update({documento: True});
                    break;
                else:
                    dictDocumentoOsVarias.update({documento: False});
            
        for linha in range(0, len(tabelaTop2Filtrada["NUM. DOCUMENTO"])):
            for documento in dictDocumentoContFiltrado:
                if tabelaTop2Filtrada.loc[linha, "NUM. DOCUMENTO"] == documento:
                    if dictDocumentoOsVarias[documento]:
                        tabelaTop2Filtrada.at[linha, "OS"] = "Várias"
                    if dictDocumentoPrefixoVarios[documento]:
                        tabelaTop2Filtrada.at[linha, "PREFIXO"] = "Vários"

        tabelaTop2Filtrada.drop_duplicates(subset=["PN", "NUM. DOCUMENTO", "QTDE"], inplace=True);
        
        with pd.ExcelWriter(wbLiberacaoCaminho, mode="a", if_sheet_exists="overlay") as writer:
            tabelaTop2Filtrada.to_excel(excel_writer=writer, sheet_name=wsTopico2Nome, index=False);
    
        return topVazio;

    #Insere a validação de dados para o problema raiz e cria os títulos dos tópicos, além de definir o número de linhas de cada tópico.
    def insereValidacaoProbRaiz(wbLiberacao, wsTopicoNome, listaTopExist, numLinhasT1, numLinhasT2, listaTopVazios):
        
        wsTopico = wbLiberacao[wsTopicoNome];
        
        wsApoio = wbLiberacao["Apoio"];

        validacao_dados = DataValidation(type="list", formula1=f"=Apoio!$A$2:$A${wsApoio.max_row}", allow_blank=True, error="Favor selecionar um problema da lista",
                                        errorTitle="Entrada Inválida", prompt="Favor selecionar um problema da lista", promptTitle="Problema Raiz", showErrorMessage=True, showInputMessage=True);

        wsTopico.add_data_validation(validacao_dados);

        fontTopicos = Font(name="Calibri", size=11, bold=True, color="ffffff");
        corTitulo = PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="006c4f"));
        
        wsTopico.insert_rows(1, 4);
        cellTitulo = wsTopico.cell(row=1, column=1, value="FOLLOW UP DE PLANEJAMENTO, COMPRAS, REPAROS E LOGÍSTICA");
        cellTitulo.font = Font(bold=True, color="3f3f3f");
        
        cellDataTitulo = wsTopico.cell(row=2, column=1, value=date.today().strftime("%d/%m/%Y"));
        cellDataTitulo.font = Font(color="3f3f3f");

        if (listaTopExist[0] and listaTopExist[1]):
            cellTitTop1 = wsTopico.cell(row=4, column=1, value="TOPICO 1 - DEMANDAS SEM SUPRIMENTOS");
            cellTitTop1.font = fontTopicos;
            cellTitTop1.fill = corTitulo;
            if (listaTopVazios[0]):
                linFimTop1 = 6;
            else:
                linFimTop1 = numLinhasT1+4;
            linIniTop2 = linFimTop1+6;
            linFimTop2 = linIniTop2+numLinhasT2-1;
            for i in range(len(listaTopExist)):
                if (numLinhasT1 > 1 and i == 0):
                    validacao_dados.add(f"O6:O{linFimTop1}");
                elif (numLinhasT2 > 1 and i == 1):
                    cellTitTop2 = wsTopico.cell(row=linIniTop2-1, column=1, value="TOPICO 2 - MATERIAL NÃO LIBERADO PARA COMPRA OU TRANSFERÊNCIA");
                    cellTitTop2.font = fontTopicos;
                    cellTitTop2.fill = corTitulo;
                    validacao_dados.add(f"R{linIniTop2+1}:R{linFimTop2}");
        else:
            if listaTopExist[0]:
                linFimTop1 = wsTopico.max_row;
                linIniTop2 = 0;
                linFimTop2 = 0;
                cellTitTop1 = wsTopico.cell(row=4, column=1, value="TOPICO 1 - DEMANDAS SEM SUPRIMENTOS");
                cellTitTop1.font = fontTopicos;
                cellTitTop1.fill = corTitulo;
                validacao_dados.add(f"O6:O{wsTopico.max_row}");
                
                
            elif listaTopExist[1]:
                linIniTop2 = 5;
                linFimTop2 = wsTopico.max_row;
                linFimTop1 = 0;
                cellTitTop2 = wsTopico.cell(row=4, column=1, value="TOPICO 2 - MATERIAL NÃO LIBERADO PARA COMPRA OU TRANSFERÊNCIA");
                cellTitTop2.font = fontTopicos;
                cellTitTop2.fill = corTitulo;
                validacao_dados.add(f"R6:R{wsTopico.max_row}");
        
        wbLiberacao.save(wbLiberacaoCaminho);
        wbLiberacao.close();
        return linFimTop1, linIniTop2, linFimTop2;

    def uneTopicos(wbLiberacao, wsTopico1Nome, wsTopico2Nome):
        wsTopico1 = wbLiberacao[wsTopico1Nome];
        wsTopico2 = wbLiberacao[wsTopico2Nome];

        numLinhasT1 = wsTopico1.max_row;
        numLinhasT2 = wsTopico2.max_row;

        if listaTopVazios[0]:
            espacoTopicos = 7; 
        else:
            espacoTopicos = 6;
        
        linhaInsTop2 = espacoTopicos + numLinhasT1;

        for linha in range(1, numLinhasT2 + 1):
            for coluna in range(1, wsTopico2.max_column + 1):
                wsTopico1.cell(linhaInsTop2 + linha-1, coluna).value = wsTopico2.cell(linha, coluna).value;

        wsTopico1.title = dataHoje;
        wbLiberacao.remove(wsTopico2);
        wbLiberacao.save(wbLiberacaoCaminho);
        return numLinhasT1, numLinhasT2;

    #Aplica o tamanho das colunas, as cores dos tipos de sugestão, cria as tabelas e corrige o alinhamento.
    def aplicaEstilo(wbLiberacao, wsTopicoNome, linFimTop1, linIniTop2, linFimTop2, listaTopExist):
        wsTopico = wbLiberacao[wsTopicoNome];
        estiloTabela = TableStyleInfo(name="TableStyleMedium19", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False);
        fillEntrega = PatternFill(patternType="solid", fgColor=Color(rgb="edf3af"), fill_type="solid");
        fillOrdem = PatternFill(patternType="solid", fgColor=Color(rgb="a2d7f0"), fill_type="solid");
        colDatasT1 = [12, 13];
        colDatasT2 = [9, 16];
        linCabecalhos = [5, linIniTop2];
        dtHoje = date.today()
        diaTab = dtHoje.strftime("%d");
        mesTab = dtHoje.strftime("%b").capitalize();
        anoTab = dtHoje.strftime("%Y");

        for topExist in range(len(listaTopExist)):
            if (listaTopExist[topExist] and topExist == 0):
                
                wsTopico.add_table(Table(displayName=f"Tópico1_{diaTab}_{mesTab}_{anoTab}", ref=f"A5:T{linFimTop1}", tableStyleInfo=estiloTabela));
                for linha in range(6, linFimTop1+1):
                    for coluna in colDatasT1:
                        wsTopico.cell(linha, coluna).number_format = "dd/mmm/yy";
            
                for coluna in range(1, 21):
                        wsTopico.cell(linCabecalhos[0], coluna).font = Font(color=Color(rgb="000000"));
            
            if (listaTopExist[topExist] and topExist == 1):
                wsTopico.add_table(Table(displayName=f"Tópico2_{diaTab}_{mesTab}_{anoTab}", ref=f"A{linIniTop2}:W{linFimTop2}", tableStyleInfo=estiloTabela));
                for linha in range(linIniTop2+1, linFimTop2+1):
                    celula = wsTopico.cell(linha, 6);
                    if celula.value == "Entrega inbound planejada": 
                        celula.fill = fillEntrega;
                    elif celula.value == "Ordem planejada":
                        celulaOrgOrigem = wsTopico.cell(linha, 7)
                        if celulaOrgOrigem.value != "":
                            celulaOrgOrigem.value = "";
                        celula.fill = fillOrdem;
                
                for linha in range(linIniTop2+1, linFimTop2+1):
                    wsTopico.cell(linha, 17).alignment = Alignment(vertical="bottom", horizontal="left");
                    for coluna in colDatasT2:
                        wsTopico.cell(linha, coluna).number_format = "dd/mmm/yy";
                
                for coluna in range(1, 24):
                        wsTopico.cell(linCabecalhos[1], coluna).font = Font(color=Color(rgb="000000"));
        
        for coluna in range(1, wsTopico.max_column+1):
            for linha in range(1, wsTopico.max_row+1):
                    if ((linha >= 6 and linha <= linFimTop1) and (coluna == 14 or coluna == 20) and listaTopExist[0]):
                        wsTopico.cell(linha, coluna).alignment = Alignment(vertical="bottom", horizontal="fill");
                    elif ((linha >= linIniTop2+1 and linha <= linFimTop2) and (coluna == 17 or coluna == 23) and listaTopExist[1]):
                        wsTopico.cell(linha, coluna).alignment = Alignment(vertical="bottom", horizontal="fill");
                    else:
                        wsTopico.cell(linha, coluna).alignment = Alignment(vertical="bottom", horizontal="left");
    
        tamColunas = [17, 9, 5, 5, 5, 15, 10, 10, 12, 7, 8, 10, 10, 6, 12, 10, 16, 11, 11, 11 ,11];

        dim_holder = DimensionHolder(worksheet=wsTopico);

        for col in range(1, 22):
            dim_holder[get_column_letter(col)] = ColumnDimension(wsTopico, min=col, max=col, width=tamColunas[col-1]);

        wsTopico.column_dimensions = dim_holder;

        wsTopico.sheet_view.showGridLines = False;

        wbLiberacao.save(wbLiberacaoCaminho);
    
    #Se o dia anterior for sábado ou domingo, informa o dia de sexta, do contrário pega o dia anterior.
    def calculaDiaAnterior():
        diaOntem = date.today()-pd.Timedelta(days=1);
        if date.weekday(diaOntem) == 6:
            return (diaOntem-pd.Timedelta(days=2)).strftime("%d-%m");
        else:
            return diaOntem.strftime("%d-%m");
        

    #Busca as informações Ação, Problema Raiz, OBS, PN s/C, Condição e Descrição dos tópicos do dia anterior e insere na planilha do dia atual.
    #Busca também as informações de restrição dos itens que estão na planilha de restrição.
    def buscarInformacoesAnteriores(wbLiberacao, wsTopicoNome, wbItensRestricaoCaminho, linFimTop1, linIniTop2, linFimTop2, listaTopVazios):
        wbItensRestricao = load_workbook(wbItensRestricaoCaminho);
        wsItensRestricao = wbItensRestricao[wbItensRestricao.sheetnames[0]]
        wsTopico = wbLiberacao[wsTopicoNome];
        wsTopicoDiaAnteriorNome = calculaDiaAnterior()
        if not wsTopicoDiaAnteriorNome in wbLiberacao.sheetnames:
            if wsTopicoDiaAnteriorNome + " T1" in wbLiberacao.sheetnames:
                wsTopicoDiaAnteriorNome = wsTopicoDiaAnteriorNome + "T1";
            elif wsTopicoDiaAnteriorNome + " T2" in wbLiberacao.sheetnames:
                wsTopicoDiaAnteriorNome = wsTopicoDiaAnteriorNome + "T2";
        wsTopicoDiaAnterior = wbLiberacao[wsTopicoDiaAnteriorNome];
        linFimT1Old = 0;
        linIniT2Old = 0;
        linFimT2Old = 0;
        fonteRestricao = Font(bold=True, color="3f3f3f");

        for linha in range(6, wsTopicoDiaAnterior.max_row+1):
            linhaTopOld = wsTopicoDiaAnterior.cell(linha, 1).value;
            if (linhaTopOld == "" and not listaTopVazios[0] and linFimT1Old == 0):
                linFimT1Old = linha-1;
            if (linhaTopOld == "PN" and not listaTopVazios[0] and linFimT1Old == 0):
                linFimT1Old = linha-6;
            if (linhaTopOld == "PN" and not listaTopVazios[1] and linIniT2Old == 0):
                linIniT2Old = linha+1;
            if (linIniT2Old != 0):
                linFimT2Old = wsTopicoDiaAnterior.max_row+1;
                break;
        
        if (not listaTopVazios[0]):
            for linhaNew in range(6, linFimTop1):
                linConcatNew = wsTopico.cell(linhaNew, 14).value;
                linPnNew = wsTopico.cell(linhaNew, 6).value;
                for linhaNumRestricao in range(2, wsItensRestricao.max_row):
                    linPnRestricao = wsItensRestricao.cell(linhaNumRestricao, 1).value;
                    if (linPnNew == linPnRestricao):
                        wsTopico.cell(linhaNew, 6).font = fonteRestricao;
                        wsTopico.cell(linhaNew, 17).value = wsItensRestricao.cell(linhaNumRestricao, 2).value;
                        wsTopico.cell(linhaNew, 18).font = fonteRestricao;
                        wsTopico.cell(linhaNew, 19).font = fonteRestricao;
                
                for linhaOld in range(6, linFimT1Old):
                    linConcatOld = wsTopicoDiaAnterior.cell(linhaOld, 14).value;
                    if (linConcatOld == linConcatNew):
                        for coluna in range(15, 18):
                            wsTopico.cell(linhaNew, coluna).value = wsTopicoDiaAnterior.cell(linhaOld, coluna).value;
                        if (wsTopicoDiaAnterior.cell(linhaOld, 18).comment is not None):
                            wsTopico.cell(linhaNew, 18).comment = wsTopicoDiaAnterior.cell(linhaOld, 18).comment;
                            wsTopico.cell(linhaNew, 18).comment.text = f"Aba {wsTopicoDiaAnteriorNome}: {wsTopicoDiaAnterior.cell(linhaOld, 15).value}\n" + wsTopico.cell(linhaNew, 18).comment.text;
                        else:
                            if (wsTopicoDiaAnterior.cell(linhaOld, 15).value != ""):
                                wsTopico.cell(linhaNew, 18).comment = Comment(f"Aba {wsTopicoDiaAnteriorNome}: {wsTopicoDiaAnterior.cell(linhaOld, 15).value}", "");
        wbItensRestricao.close();

        if (not listaTopVazios[1]):
            contLinhas = 0;
            for linhaNew in range(linIniTop2, linFimTop2):
                linConcatNew = wsTopico.cell(linhaNew, 17).value;
                linPnNew = wsTopico.cell(linhaNew, 1).value;
                for linhaNumRestricao in range(2, wsItensRestricao.max_row):
                    linPnRestricao = wsItensRestricao.cell(linhaNumRestricao, 1).value
                    if (linPnNew == linPnRestricao):
                        wsTopico.cell(linhaNew, 1).font = fonteRestricao;
                        wsTopico.cell(linhaNew, 20).value = wsItensRestricao.cell(linhaNumRestricao, 2).value;
                        wsTopico.cell(linhaNew, 21).font = fonteRestricao;
                        wsTopico.cell(linhaNew, 22).font = fonteRestricao;
                    
                for linhaOld in range(linIniT2Old, linFimT2Old):
                    contLinhas += 1;
                    linConcatOld = wsTopicoDiaAnterior.cell(linhaOld, 17).value
                    if (linConcatOld == linConcatNew):
                        for coluna in range(18, 21):
                            wsTopico.cell(linhaNew, coluna).value = wsTopicoDiaAnterior.cell(linhaOld, coluna).value;
                        if (wsTopicoDiaAnterior.cell(linhaOld, 21).comment is not None):
                            wsTopico.cell(linhaNew, 21).comment = wsTopicoDiaAnterior.cell(linhaOld, 21).comment;
                            wsTopico.cell(linhaNew, 21).comment.text = f"Aba {wsTopicoDiaAnteriorNome}: {wsTopicoDiaAnterior.cell(linhaOld, 18).value}\n" + wsTopico.cell(linhaNew, 21).comment.text;
                        else:
                            if (wsTopicoDiaAnterior.cell(linhaOld, 18).value != ""):
                                wsTopico.cell(linhaNew, 21).comment = Comment(f"Aba {wsTopicoDiaAnteriorNome}: {wsTopicoDiaAnterior.cell(linhaOld, 18).value}", "");
        wbItensRestricao.close();
        wbLiberacao.save(wbLiberacaoCaminho);
    
    if ent_caminho_pasta_liberacao.get() == "":
        messagebox.showerror("Caminho inválido", "Favor informar o caminho da pasta de trabalho da liberação.")
        return
    criarBackupPastaLiberacao()
    #Se os dois tópicos existirem, filtra os dados de cada tópico, insere a validação de dados, aplica o estilo e busca as informações anteriores.
    if (listaTopExist[0] and listaTopExist[1]):
        #Se a planilha de liberação estiver aberta, exibe uma mensagem de erro e retorna para a tela inicial.
        try:
            listaTopVazios[0] = filtraTopico1(wbLiberacaoCaminho, wsTopico1Nome, listaTopicos);
            listaTopVazios[1] = filtraTopico2(wbLiberacaoCaminho, wsTopico2Nome, listaTopicos);
        except PermissionError:
            messagebox.showerror("Planilha de liberação aberta", "A planilha de liberação está aberta pelo usuário, impossibilitando a formatação, favor fechar e tentar novamente.");
            return;
        wbLiberacao = load_workbook(wbLiberacaoCaminho)
        numLinhasT1, numLinhasT2 = uneTopicos(wbLiberacao, wsTopico1Nome, wsTopico2Nome);
        linFimTop1, linIniTop2, linFimTop2 = insereValidacaoProbRaiz(wbLiberacao, dataHoje, listaTopExist, numLinhasT1, numLinhasT2, listaTopVazios);
        aplicaEstilo(wbLiberacao, dataHoje, linFimTop1, linIniTop2, linFimTop2, listaTopExist);
        #Para fazer alterações na planilha de itens com restrições, seguir o padrão dentro da planilha se não o programa não irá funcionar corretamente.
        buscarInformacoesAnteriores(wbLiberacao, dataHoje, caminho_Itens_Restricao, linFimTop1, linIniTop2, linFimTop2, listaTopVazios)
    #Se apenas o tópico 1 existir, filtra os dados do tópico 1, insere a validação de dados, aplica o estilo e busca as informações anteriores.    
    else:
        print("Existe apenas um tópico")
        i = 1;
        while i <= len(listaTopicos):
            topicoIdx = i-1;
            try:
                if listaTopExist[topicoIdx] and i == 1:

                    listaTopVazios.append(filtraTopico1(wbLiberacaoCaminho, wsTopico1Nome, listaTopicos));
                    wsTopicoNome = wsTopico1Nome;

                elif listaTopExist[topicoIdx] and i == 2:

                    listaTopVazios.append(filtraTopico2(wbLiberacaoCaminho, wsTopico2Nome, listaTopicos));
                    wsTopicoNome = wsTopico2Nome;
            except PermissionError:
                messagebox.showerror("Planilha de liberação aberta", "A planilha de liberação está aberta pelo usuário, impossibilitando a formatação, favor fechar e tentar novamente.");
                return
            i +=1;
        wbLiberacao = load_workbook(wbLiberacaoCaminho)
        linFimTop1, linIniTop2, linFimTop2 = insereValidacaoProbRaiz(wbLiberacao, wsTopicoNome, listaTopExist, numLinhasT1, numLinhasT2, listaTopVazios);
        aplicaEstilo(wbLiberacao, wsTopicoNome, linFimTop1, linIniTop2, linFimTop2, listaTopExist);
        buscarInformacoesAnteriores(wbLiberacao, wsTopicoNome, caminho_Itens_Restricao, linFimTop1, linIniTop2, linFimTop2, listaTopVazios);

    wbLiberacao.close()
    messagebox.showinfo("Conclusão", "Formatação concluída com sucesso!");

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl");

bgCor = "#eef1e6";

if Path("flag-icon.ico").exists() == False:
    messagebox.showerror("Erro", "O arquivo 'flag-icon.ico' não foi encontrado. Favor verificar o diretório do programa.");
    exit();
elif Path("logo_lider.png").exists() == False:
    messagebox.showerror("Erro", "O arquivo 'logo_lider.png' não foi encontrado. Favor verificar o diretório do programa.");
    exit();

janela = tk.Tk();
janela.title("Formatação Follow-Up");
janela.iconbitmap(Path("flag-icon.ico"));
janela.config(bg=bgCor);
janela.minsize(940, 540);

#Função para selecionar o caminho dos tópicos.
def selecionar_caminho_topicos():
    caminho_topicos = filedialog.askdirectory(title="Selecione um caminho");
    ent_caminho_topicos.delete(0, tk.END);
    ent_caminho_topicos.insert(0, caminho_topicos);

#Função para escolher a pasta de trabalho da liberação.
def selecionar_pasta_liberacao():
    caminho_liberacao = filedialog.askopenfilename(title="Selecione um arquivo");
    ent_caminho_pasta_liberacao.delete(0, tk.END);
    ent_caminho_pasta_liberacao.insert(0, caminho_liberacao);

#Função para escolher a pasta de trabalho dos itens com restrição de liberação.
def selecionar_pasta_restricao():
    caminho_restricao = filedialog.askopenfilename(title="Selecione um arquivo");
    ent_caminho_pasta_restricao.delete(0, tk.END);
    ent_caminho_pasta_restricao.insert(0, caminho_restricao);

def centralizar_Janela(janela):
    altura_janela = 540;
    largura_janela = 940;

    largura_tela = janela.winfo_screenwidth();
    altura_tela = janela.winfo_screenheight();

    x = (largura_tela - largura_janela) // 2;
    y = (altura_tela - altura_janela) // 2;

    janela.geometry(f"{largura_janela}x{altura_janela}+{x}+{y}");

#Cria as linhas e colunas da janela principal.
frames = {"A0":tk.Frame(), "A1":tk.Frame(), "A2":tk.Frame(), "A3":tk.Frame(), "A4":tk.Frame(), "A5":tk.Frame(), "A6":tk.Frame(), "A7":tk.Frame(), "A8":tk.Frame(), "A9":tk.Frame(),
          "B0":tk.Frame(), "B1":tk.Frame(), "B2":tk.Frame(), "B3":tk.Frame(), "B4":tk.Frame(), "B5":tk.Frame(), "B6":tk.Frame(), "B7":tk.Frame(), "B8":tk.Frame(), "B9":tk.Frame(),
          "C0":tk.Frame(), "C1":tk.Frame(), "C2":tk.Frame(), "C3":tk.Frame(), "C4":tk.Frame(), "C5":tk.Frame(), "C6":tk.Frame(), "C7":tk.Frame(), "C8":tk.Frame(), "C9":tk.Frame()};

listaColunas = ["A", "B", "C"]

for coluna in range(0, 3):
    janela.columnconfigure(coluna, weight=1, minsize=50);
    for linha in range(0, 9):
        frames[listaColunas[coluna] + str(linha)].grid(row=linha, column=coluna, padx=10, pady=10);
        janela.rowconfigure(linha, weight=1, minsize=25);

fontLblEntries = ["Calibri", 12];

#Cria os textos, botões, entradas e imagens da janela principal.

logo_lider = ImageTk.PhotoImage(Image.open(Path("logo_lider.png")).resize((234, 38)));

lbl_logoLider = tk.Label(master=frames["A0"], image=logo_lider, bg=bgCor);
lbl_logoLider.pack();
lbl_tituloFup = tk.Label(master=frames["B0"], text="Formatação do Follow-Up", font=("Arial", 21, "bold"), fg="#00513a", bg=bgCor);
lbl_tituloFup.pack();

lbl_caminho_topicos = tk.Label(master=frames["A1"], text="Caminho das planilhas dos tópicos: ", bg=bgCor, font=fontLblEntries);
lbl_caminho_topicos.pack(padx=10);
ent_caminho_topicos = tk.Entry(master=frames["B1"], width=50, font=fontLblEntries);
ent_caminho_topicos.pack();
btn_caminho_topicos = tk.Button(master=frames["C1"], width=20, text="Selecionar caminho", command=selecionar_caminho_topicos, font=fontLblEntries);
btn_caminho_topicos.pack(padx=10);

lbl_caminho_pasta_liberacao = tk.Label(master=frames["A2"], text="Caminho da pasta de trabalho da liberação: ", bg=bgCor, font=fontLblEntries);
lbl_caminho_pasta_liberacao.pack(padx=10);
ent_caminho_pasta_liberacao = tk.Entry(master=frames["B2"], width=50, font=fontLblEntries);
ent_caminho_pasta_liberacao.pack();
btn_caminho_pasta_liberacao = tk.Button(master=frames["C2"], width=20, text="Selecionar arquivo", command=selecionar_pasta_liberacao, font=fontLblEntries);
btn_caminho_pasta_liberacao.pack(padx=10);

lbl_caminho_pasta_restricao = tk.Label(master=frames["A3"], text="Caminho da pasta de trabalho dos itens " + "\n que possuem restrição de liberação: ", bg=bgCor, font=fontLblEntries);
lbl_caminho_pasta_restricao.pack(padx=10);
ent_caminho_pasta_restricao = tk.Entry(master=frames["B3"], width=50, font=fontLblEntries);
ent_caminho_pasta_restricao.pack();
btn_caminho_pasta_restricao = tk.Button(master=frames["C3"], width=20, text="Selecionar arquivo", command=selecionar_pasta_restricao, font=fontLblEntries);
btn_caminho_pasta_restricao.pack(padx=10);

lbl_undNegoc = tk.Label(master=frames["A4"], text="Unidade de Negócio:", width=20, bg=bgCor, font=fontLblEntries);
lbl_undNegoc.pack();
listaUndNegoc = ["HELICOPTEROS", "MANUTENCAO"];
combo_undNegoc = ttk.Combobox(master=frames["B4"], values=listaUndNegoc, font=fontLblEntries, justify="center");
combo_undNegoc.pack();

lbl_entregaInbound = tk.Label(master=frames["A5"], text="Dias de análise\n Entrega Inbound Planejada:", width=30, bg=bgCor, font=fontLblEntries);
lbl_entregaInbound.pack();
ent_entregaInbound = tk.Entry(master=frames["B5"], width=10, justify="center", font=fontLblEntries);
ent_entregaInbound.pack();

lbl_ordemPlanejada = tk.Label(master=frames["A6"], text="Dias de análise\n Ordem Planejada Geral:", width=30, bg=bgCor, font=fontLblEntries);
lbl_ordemPlanejada.pack();
ent_ordemPlanejada = tk.Entry(master=frames["B6"], width=10, justify="center", font=fontLblEntries);
ent_ordemPlanejada.pack();

lbl_ordemPlanejadaCDB = tk.Label(master=frames["A7"], text="Dias de análise\n Ordem Planejada CDB-UOH:", width=30, bg=bgCor, font=fontLblEntries);
lbl_ordemPlanejadaCDB.pack();
ent_ordemPlanejadaCDB = tk.Entry(master=frames["B7"], width=10, justify="center", font=fontLblEntries);
ent_ordemPlanejadaCDB.pack();

btn_formatar = tk.Button(master=frames["B8"], width=20, text="Formatar", command=formatar, font=fontLblEntries);
btn_formatar.pack(pady=10);

centralizar_Janela(janela);

#Carrega os filtros salvos, se existirem, e preenche os campos correspondentes.
caminho_Filtros = Path("filtros.txt");
if caminho_Filtros.exists():
    arqFiltros = open(caminho_Filtros, "r");
    listaFiltros = arqFiltros.readlines();
    if len(listaFiltros) < 7:
        messagebox.showerror("Filtros inválidos", "O arquivo de filtros está incompleto ou corrompido. Favor verificar o arquivo.");
        arqFiltros.close();
    else:
        for filtro in range(len(listaFiltros)):
            listaFiltros[filtro] = listaFiltros[filtro].strip();
        ent_caminho_topicos.insert(0, listaFiltros[0]);
        ent_caminho_pasta_liberacao.insert(0, listaFiltros[1]);
        ent_caminho_pasta_restricao.insert(0, listaFiltros[2])
        combo_undNegoc.insert(0, listaFiltros[3].strip());
        ent_entregaInbound.insert(0, listaFiltros[4]);
        ent_ordemPlanejada.insert(0, listaFiltros[5]);
        ent_ordemPlanejadaCDB.insert(0, listaFiltros[6]);
        arqFiltros.close()

#Inicia a execução da janela principal.
janela.mainloop();
        


